import sqlite3
import openpyxl
import re
import os

DB_PATH = '/root/clinic-app/data/clinic.db'
IMPORT_DIR = '/root/clinic-app/data/import_excel'
YEAR = 2026

conn = sqlite3.connect(DB_PATH)
conn.row_factory = sqlite3.Row
conn.execute("PRAGMA foreign_keys = ON")
conn.executescript("""
    CREATE TABLE IF NOT EXISTS patients (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL UNIQUE,
        phone TEXT DEFAULT '',
        age TEXT DEFAULT '',
        gender TEXT DEFAULT '',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );
    CREATE TABLE IF NOT EXISTS appointments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        patient_id INTEGER NOT NULL,
        year INTEGER NOT NULL,
        month INTEGER NOT NULL,
        day INTEGER NOT NULL,
        time_slot TEXT DEFAULT '',
        project TEXT NOT NULL,
        amount INTEGER DEFAULT 0,
        remark TEXT DEFAULT '',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (patient_id) REFERENCES patients(id) ON DELETE CASCADE
    );
    CREATE TABLE IF NOT EXISTS consumption (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        patient_id INTEGER NOT NULL,
        month INTEGER NOT NULL,
        day INTEGER NOT NULL,
        time_slot TEXT DEFAULT '',
        project TEXT NOT NULL,
        amount INTEGER DEFAULT 0,
        sync_to_apt INTEGER DEFAULT 0,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (patient_id) REFERENCES patients(id) ON DELETE CASCADE
    );
    CREATE TABLE IF NOT EXISTS appointment_consumption_link (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        appointment_id INTEGER NOT NULL,
        consumption_id INTEGER NOT NULL,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (appointment_id) REFERENCES appointments(id) ON DELETE CASCADE,
        FOREIGN KEY (consumption_id) REFERENCES consumption(id) ON DELETE CASCADE,
        UNIQUE(appointment_id, consumption_id)
    );
    CREATE TABLE IF NOT EXISTS operation_history (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        action_type TEXT NOT NULL,
        target_type TEXT NOT NULL,
        target_id INTEGER,
        data TEXT NOT NULL,
        description TEXT NOT NULL,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );
    CREATE TABLE IF NOT EXISTS treatment_projects (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL UNIQUE,
        default_price INTEGER DEFAULT 0,
        sort_order INTEGER DEFAULT 0,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );
""")
for col in ['age', 'gender']:
    try:
        conn.execute(f"ALTER TABLE patients ADD COLUMN {col} TEXT DEFAULT ''")
    except Exception:
        pass
try:
    conn.execute("ALTER TABLE consumption ADD COLUMN time_slot TEXT DEFAULT ''")
except Exception:
    pass
try:
    conn.execute("ALTER TABLE consumption ADD COLUMN sync_to_apt INTEGER DEFAULT 0")
except Exception:
    pass

stats = {'files': 0, 'patients': 0, 'appointments': 0, 'consumption': 0}

for filename in sorted(os.listdir(IMPORT_DIR)):
    if not filename.endswith('.xlsx'):
        continue
    month_match = re.search(r'(\d+)月', filename)
    if not month_match:
        print(f'  SKIP {filename}: cannot parse month')
        continue
    month = int(month_match.group(1))
    filepath = os.path.join(IMPORT_DIR, filename)
    print(f'Processing {filename} (month {month})...')

    wb = openpyxl.load_workbook(filepath)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Parse date headers from row 2
        dates = {}  # col_index -> day
        for col in range(2, 30, 2):
            val = ws.cell(row=2, column=col).value
            if val:
                val = str(val).replace('\n', ' ')
                date_match = re.search(r'(\d{2})-(\d{2})', val)
                if date_match:
                    day = int(date_match.group(2))
                    dates[col] = day

        for row in range(4, 16):
            time_slot = ws.cell(row=row, column=1).value
            if not time_slot:
                continue
            time_slot = str(time_slot).strip()

            for date_col, day in dates.items():
                name_col = date_col
                remark_col = date_col + 1

                name_val = ws.cell(row=row, column=name_col).value
                if not name_val:
                    continue

                name = str(name_val).strip()
                remark_val = ws.cell(row=row, column=remark_col).value
                remark = str(remark_val).strip() if remark_val else ''

                # Parse amount from remark
                amount = 0
                amount_match = re.search(r'(\d+)\s*$', remark)
                if amount_match:
                    amount = int(amount_match.group(1))

                # Parse project from remark (remove trailing number)
                project = re.sub(r'[\d\s]+$', '', remark).strip()
                if not project:
                    project = '治疗'

                # Get or create patient
                conn.execute("INSERT OR IGNORE INTO patients (name) VALUES (?)", (name,))
                if conn.execute("SELECT changes()").fetchone()[0] > 0:
                    stats['patients'] += 1
                patient_id = conn.execute("SELECT id FROM patients WHERE name=?", (name,)).fetchone()['id']

                # Insert appointment
                conn.execute(
                    "INSERT INTO appointments (patient_id, year, month, day, time_slot, project, amount, remark) VALUES (?,?,?,?,?,?,?,?)",
                    (patient_id, YEAR, month, day, time_slot, project, amount, remark)
                )
                stats['appointments'] += 1

    wb.close()

conn.commit()
conn.close()
print(f'\nDone! Stats: {stats}')
