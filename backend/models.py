import sqlite3
import os

DB_PATH = os.path.join(os.path.dirname(__file__), 'data', 'clinic.db')

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    # 启用外键约束
    conn.execute("PRAGMA foreign_keys = ON")
    return conn

def init_db():
    conn = get_db()
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS patients (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            phone TEXT DEFAULT '',
            age TEXT DEFAULT '',
            gender TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS appointments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            patient_id INTEGER NOT NULL,
            year INTEGER NOT NULL,
            month INTEGER NOT NULL,
            day INTEGER NOT NULL,
            time_slot TEXT DEFAULT '',
            project TEXT NOT NULL,
            amount INTEGER DEFAULT 0,
            remark TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (patient_id) REFERENCES patients(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS consumption (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            patient_id INTEGER NOT NULL,
            month INTEGER NOT NULL,
            day INTEGER NOT NULL,
            time_slot TEXT DEFAULT '',
            project TEXT NOT NULL,
            amount INTEGER DEFAULT 0,
            sync_to_apt INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (patient_id) REFERENCES patients(id) ON DELETE CASCADE
        );
        
        -- 预约与耗卡的关联表（替代 remark 中的 #id）
        CREATE TABLE IF NOT EXISTS appointment_consumption_link (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            appointment_id INTEGER NOT NULL,
            consumption_id INTEGER NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (appointment_id) REFERENCES appointments(id) ON DELETE CASCADE,
            FOREIGN KEY (consumption_id) REFERENCES consumption(id) ON DELETE CASCADE,
            UNIQUE(appointment_id, consumption_id)
        );

        CREATE INDEX IF NOT EXISTS idx_appointments_date ON appointments(year, month);
        CREATE INDEX IF NOT EXISTS idx_appointments_patient ON appointments(patient_id);
        CREATE INDEX IF NOT EXISTS idx_consumption_patient ON consumption(patient_id);
        CREATE INDEX IF NOT EXISTS idx_consumption_date ON consumption(month, day);

        CREATE TABLE IF NOT EXISTS operation_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            action_type TEXT NOT NULL,
            target_type TEXT NOT NULL,
            target_id INTEGER,
            data TEXT NOT NULL,
            description TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS treatment_projects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            default_price INTEGER DEFAULT 0,
            sort_order INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    ''')
    # 迁移：为已有数据库添加 age/gender/time_slot 列
    for col in ['age', 'gender']:
        try:
            conn.execute(f"ALTER TABLE patients ADD COLUMN {col} TEXT DEFAULT ''")
        except Exception:
            pass
    try:
        conn.execute("ALTER TABLE consumption ADD COLUMN time_slot TEXT DEFAULT ''")
    except Exception:
        pass
    try:
        conn.execute("ALTER TABLE consumption ADD COLUMN sync_to_apt INTEGER DEFAULT 0")
    except Exception:
        pass
    # 初始化默认治疗项目
    existing = conn.execute("SELECT COUNT(*) FROM treatment_projects").fetchone()[0]
    if existing == 0:
        defaults = [
            ('\u5eb7\u590d', 400, 1),
            ('\u653e\u677e', 169, 2),
            ('\u8bc4\u4f30', 99, 3),
        ]
        for name, price, sort in defaults:
            conn.execute("INSERT OR IGNORE INTO treatment_projects (name, default_price, sort_order) VALUES (?,?,?)",
                        (name, price, sort))
    conn.commit()
    # 迁移：将 remark 中的 #id 迁移到关联表
    try:
        cursor = conn.cursor()
        # 查找所有包含 #id 的 remark
        cursor.execute("SELECT id, remark FROM appointments WHERE remark LIKE '%#%'")
        for row in cursor.fetchall():
            apt_id = row[0]
            remark = row[1]
            match = __import__('re').search(r'#(\d+)$', remark)
            if match:
                con_id = int(match.group(1))
                # 检查关联是否已存在
                cursor.execute("SELECT id FROM appointment_consumption_link WHERE appointment_id=? AND consumption_id=?", (apt_id, con_id))
                if not cursor.fetchone():
                    # 检查 consumption 是否存在
                    cursor.execute("SELECT id FROM consumption WHERE id=?", (con_id,))
                    if cursor.fetchone():
                        cursor.execute("INSERT INTO appointment_consumption_link (appointment_id, consumption_id) VALUES (?, ?)", (apt_id, con_id))
        conn.commit()
    except Exception as e:
        print(f"迁移 remark #id 时出错: {e}")
    
    conn.commit()
    conn.close()

def import_from_excel():
    import openpyxl
    from datetime import datetime
    
    try:
        base_dir = r'C:\Users\26259\Desktop'
        app_dir = os.path.join(base_dir, '预约', '2026预约表')
    except:
        app_dir = os.path.join(os.path.dirname(__file__), 'data', 'appointments')
    
    if not os.path.exists(app_dir):
        return {'error': '预约表目录不存在'}
    
    conn = get_db()
    cursor = conn.cursor()
    
    stats = {'patients_added': 0, 'appointments_added': 0}
    
    try:
        for filename in os.listdir(app_dir):
            if not filename.endswith('.xlsx'):
                continue
            
            import re
            month_match = re.search(r'(\d+)\u6708', filename)
            if not month_match:
                continue
            
            month = int(month_match.group(1))
            filepath = os.path.join(app_dir, filename)
            
            wb = openpyxl.load_workbook(filepath)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                header = []
                for col in range(2, 30):
                    val = ws.cell(row=2, column=col).value
                    header.append(str(val) if val else '')
                
                for row in range(4, 18):
                    time_slot = ws.cell(row=row, column=1).value
                    if not time_slot:
                        continue
                    time_slot = str(time_slot)
                    
                    for col_idx in range(0, len(header), 2):
                        if col_idx + 2 > len(header):
                            break
                        
                        name_cell = ws.cell(row=row, column=col_idx + 2)
                        remark_cell = ws.cell(row=row, column=col_idx + 3)
                        
                        if not name_cell.value:
                            continue
                        
                        name = str(name_cell.value).strip()
                        remark = str(remark_cell.value) if remark_cell.value else ''
                        
                        date_info = header[col_idx] if col_idx < len(header) else ''
                        day = 1
                        date_match = re.search(r'(\d{2})-(\d{2})', date_info.replace('\n', ' '))
                        if date_match:
                            day = int(date_match.group(2))
                        
                        # Parse amount and project
                        remark_clean = remark.replace('\n', ' ')
                        amount = 0
                        amount_match = re.search(r'(\d+)$', remark_clean)
                        if amount_match:
                            amount = int(amount_match.group(1))
                        
                        project = re.sub(r'[\d\.]+[Hh]?', '', remark_clean)
                        project = re.sub(r'[Hh]+$', '', project).strip()
                        if not project:
                            project = '治疗'
                        
                        # Get or create patient
                        cursor.execute("INSERT OR IGNORE INTO patients (name) VALUES (?)", (name,))
                        if cursor.rowcount > 0:
                            stats['patients_added'] += 1
                        
                        cursor.execute("SELECT id FROM patients WHERE name=?", (name,))
                        patient_id = cursor.fetchone()[0]
                        
                        # Insert appointment
                        cursor.execute(
                            "INSERT OR IGNORE INTO appointments (patient_id, year, month, day, time_slot, project, amount, remark) VALUES (?,?,?,?,?,?,?,?)",
                            (patient_id, 2026, month, day, time_slot, project, amount, remark)
                        )
                        if cursor.rowcount > 0:
                            stats['appointments_added'] += 1
            
            wb.close()
        
        conn.commit()
        return stats
    finally:
        conn.close()
