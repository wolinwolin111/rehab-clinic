import os, re, time, threading
from collections import defaultdict
import openpyxl

try:
    APPT_FOLDER = os.path.join(
        os.environ['USERPROFILE'], 'Desktop',
        '预约', '2026预约表'
    )
except KeyError:
    APPT_FOLDER = os.path.join(os.path.dirname(__file__), 'data', 'appointments')

FILE_PATTERN = re.compile(r'(\d{4})\D+(\d{1,2})\D+')
HEADER_DATE_PATTERN = re.compile(r'\n(\d+)-(\d+)')
SHEET_DAY_RANGE = re.compile(r'(\d+)\D+(\d+)')

TIME_SLOT_ROWS = {
    '8:00-9:00': 4, '9:00-10:00': 5,
    '10:00-11:00': 6, '11:00-12:00': 7, '12:00-13:00': 8,
    '13:00-14:00': 9, '14:00-15:00': 10, '15:00-16:00': 11,
    '16:00-17:00': 12, '17:00-18:00': 13, '18:00-19:00': 14,
    '19:00-20:00': 15, '20:00-21:00': 16, '21:00-22:00': 17,
}

# 异步同步队列
_sync_queue = []
_queue_lock = threading.Lock()
_worker_running = False
_MAX_QUEUE_SIZE = 500  # prevent unbounded memory growth

def _worker():
    """后台工作线程，处理同步队列"""
    global _worker_running
    while True:
        time.sleep(2)
        with _queue_lock:
            if not _sync_queue:
                continue
            # Limit how many tasks to process per batch
            task = _sync_queue.pop(0)
        
        # 执行同步，最多重试3次
        for attempt in range(3):
            try:
                if task['type'] == 'upsert':
                    result = _sync_upsert_internal(
                        task['year'], task['month'], task['day'], task['time_slot'],
                        task['patient_name'], task.get('project', ''), 
                        task.get('amount', 0), task.get('remark', '')
                    )
                elif task['type'] == 'delete':
                    result = _sync_delete_internal(
                        task['year'], task['month'], task['day'], task['time_slot']
                    )
                elif task['type'] == 'full':
                    result = _sync_all_internal(task['appointments'])
                
                if result:
                    print(f"[ExcelSync] 任务成功: {task}")
                    break
            except Exception as e:
                print(f"[ExcelSync] 重试 {attempt + 1}/3: {e}")
                time.sleep(1)
        
        if attempt == 2:
            print(f"[ExcelSync] 任务失败: {task}")

def _sync_upsert_internal(year, month, day, time_slot, patient_name, project='', amount=0, remark=''):
    fpath = _find_file(year, month)
    if not fpath:
        return False
    try:
        wb = openpyxl.load_workbook(fpath)
    except Exception:
        return False
    ws = _find_sheet(wb, day)
    if not ws:
        wb.close(); return False
    col = _find_column(ws, day)
    if not col:
        wb.close(); return False
    row = TIME_SLOT_ROWS.get(time_slot)
    if not row:
        wb.close(); return False

    ws.cell(row=row, column=col).value = patient_name
    remark_col = col + 1
    parts = [project] if project else []
    if amount:
        parts.append(str(amount))
    ws.cell(row=row, column=remark_col).value = '\n'.join(parts) if parts else None

    try:
        wb.save(fpath)
        wb.close()
        return True
    except Exception:
        wb.close()
        return False

def _sync_delete_internal(year, month, day, time_slot):
    fpath = _find_file(year, month)
    if not fpath:
        return False
    try:
        wb = openpyxl.load_workbook(fpath)
    except Exception:
        return False
    ws = _find_sheet(wb, day)
    if not ws:
        wb.close(); return False
    col = _find_column(ws, day)
    if not col:
        wb.close(); return False
    row = TIME_SLOT_ROWS.get(time_slot)
    if not row:
        wb.close(); return False
    
    ws.cell(row=row, column=col).value = None
    ws.cell(row=row, column=col + 1).value = None
    try:
        wb.save(fpath)
        wb.close()
        return True
    except Exception:
        wb.close()
        return False

def _sync_all_internal(appointments):
    """同步所有预约到 Excel（批量操作）"""
    by_month = defaultdict(list)
    for a in appointments:
        by_month[(a['year'], a['month'])].append(a)
    
    ok = 0
    fail = 0
    errors = []
    
    for (year, month), apts in by_month.items():
        fpath = _find_file(year, month)
        if not fpath:
            errors.append(f"No file for {year}-{month}")
            fail += len(apts)
            continue
        try:
            wb = openpyxl.load_workbook(fpath)
        except Exception:
            errors.append(f"Cannot open file for {year}-{month}")
            fail += len(apts)
            continue
        
        for apt in apts:
            ws = _find_sheet(wb, apt['day'])
            if not ws:
                errors.append(f"No sheet for day {apt['day']}")
                fail += 1
                continue
            col = _find_column(ws, apt['day'])
            if not col:
                errors.append(f"No column for day {apt['day']}")
                fail += 1
                continue
            row = TIME_SLOT_ROWS.get(apt['time_slot'])
            if not row:
                errors.append(f"Invalid time_slot {apt['time_slot']}")
                fail += 1
                continue
            
            ws.cell(row=row, column=col).value = apt['patient_name']
            parts = [apt['project']] if apt.get('project') else []
            if apt.get('amount'):
                parts.append(str(apt['amount']))
            ws.cell(row=row, column=col + 1).value = '\n'.join(parts) if parts else None
            ok += 1
        
        try:
            wb.save(fpath)
        except Exception as e:
            errors.append(f"Failed to save {year}-{month}: {e}")
            ok = 0
            fail += len(apts)
        
        wb.close()
    
    return (ok, fail, errors)

# 公开接口：异步添加同步任务
def sync_upsert(year, month, day, time_slot, patient_name, project='', amount=0, remark=''):
    """异步同步预约到 Excel"""
    with _queue_lock:
        if len(_sync_queue) >= _MAX_QUEUE_SIZE:
            return  # drop task silently if queue is full
        _sync_queue.append({
            'type': 'upsert',
            'year': year, 'month': month, 'day': day, 'time_slot': time_slot,
            'patient_name': patient_name, 'project': project, 'amount': amount, 'remark': remark
        })
    _ensure_worker()


def sync_all(appointments):
    """批量同步到 Excel（异步）"""
    with _queue_lock:
        _sync_queue.append({
            'type': 'full',
            'appointments': appointments
        })
    _ensure_worker()

def _ensure_worker():
    global _worker_running
    with _queue_lock:
        if _worker_running:
            return
        _worker_running = True
    t = threading.Thread(target=_worker, daemon=True)
    t.start()

def _find_file(year, month):
    if not os.path.isdir(APPT_FOLDER):
        return None
    for f in os.listdir(APPT_FOLDER):
        if not f.endswith('.xlsx') or f.startswith('~'):
            continue
        m = FILE_PATTERN.search(f)
        if m and int(m.group(1)) == year and int(m.group(2)) == month:
            return os.path.join(APPT_FOLDER, f)
    return None

def _find_sheet(wb, day):
    for sn in wb.sheetnames:
        m = SHEET_DAY_RANGE.search(sn)
        if m and int(m.group(1)) <= day <= int(m.group(2)):
            return wb[sn]
    return None

def _find_column(ws, day):
    for col in range(2, ws.max_column + 1, 2):
        cell = ws.cell(row=2, column=col).value
        if cell:
            m = HEADER_DATE_PATTERN.search(str(cell))
            if m and int(m.group(2)) == day:
                return col
    return None

def get_queue_status():
    """获取同步队列状态"""
    with _queue_lock:
        return {'pending': len(_sync_queue), 'worker_running': _worker_running}

def clear_queue():
    """清空同步队列"""
    with _queue_lock:
        _sync_queue.clear()

def sync_delete(year, month, day, time_slot):
    fpath = _find_file(year, month)
    if not fpath:
        return False
    try:
        wb = openpyxl.load_workbook(fpath)
    except Exception:
        return False
    ws = _find_sheet(wb, day)
    if not ws:
        wb.close(); return False
    col = _find_column(ws, day)
    if not col:
        wb.close(); return False
    row = TIME_SLOT_ROWS.get(time_slot)
    if not row:
        wb.close(); return False

    ws.cell(row=row, column=col).value = None
    ws.cell(row=row, column=col + 1).value = None

    try:
        wb.save(fpath)
        wb.close()
        return True
    except Exception:
        wb.close()
        return False
