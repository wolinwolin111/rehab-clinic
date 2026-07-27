from flask import Flask, request, jsonify, send_from_directory, send_file
from flask_cors import CORS
from models import init_db, get_db, import_from_excel
from excel_sync import sync_upsert, sync_delete, sync_all
import os
import re
import time

app = Flask(__name__, static_folder='static')
CORS(app, resources={r"/api/*": {"origins": "*"}})  # public API, auth TBD
init_db()
# === Security: column name whitelist for dynamic UPDATE queries ===
_ALLOWED_COLUMNS = {
    'patients': {'name', 'phone', 'age', 'gender', 'balance', 'has_recharge'},
    'appointments': {'patient_id', 'year', 'month', 'day', 'time_slot', 'project', 'amount', 'remark'},
    'consumption': {'patient_id', 'month', 'day', 'project', 'amount', 'time_slot', 'sync_to_apt', 'recharge_amount'},
    'treatment_projects': {'name', 'default_price', 'sort_order'},
}
_BATCH_LIMIT = 200

def _deduct_balance(conn, patient_id, amount):
    """扣减患者余额，返回 True 表示成功。amount为0时直接成功"""
    if not amount or amount <= 0:
        return True
    patient = conn.execute("SELECT balance FROM patients WHERE id=?", (patient_id,)).fetchone()
    if not patient:
        return False
    if patient['balance'] < amount:
        conn.execute("UPDATE patients SET balance = 0 WHERE id=?", (patient_id,))
        return True  # 扣到0，不报错
    conn.execute("UPDATE patients SET balance = balance - ? WHERE id=?", (amount, patient_id))
    return True

def _validate_date(month, day):
    """Validate month (1-12) and day (1-31)"""
    if not (1 <= month <= 12 and 1 <= day <= 31):
        raise ValueError(f'无效日期: {month}月{day}日')

@app.route('/')
def index():
    return send_from_directory('static', 'index.html')

@app.route('/health')
def health():
    return jsonify({'status': 'ok'})

# ===== 备份与恢复 =====

@app.route('/api/backup', methods=['GET'])
def backup_db():
    """
    备份数据库：先从Excel同步到DB，然后返回数据库文件
    前端可以直接下载
    """
    import threading
    def _sync_first():
        try:
            from excel_sync import sync_all
            from models import get_db
            conn = get_db()
            rows = conn.execute(
                "SELECT a.*, p.name as patient_name FROM appointments a JOIN patients p ON a.patient_id=p.id"
            ).fetchall()
            conn.close()
            apts = [dict(r) for r in rows]
            sync_all(apts)
        except Exception:
            pass
    
    # 先异步同步Excel数据到DB
    t = threading.Thread(target=_sync_first)
    t.start()
    t.join()
    
    # 然后发送数据库文件
    from models import DB_PATH
    from flask import send_file
    return send_file(
        DB_PATH,
        mimetype='application/x-sqlite3',
        as_attachment=True,
        download_name=f'clinic_backup_{time.strftime("%Y%m%d_%H%M%S")}.db'
    )

@app.route('/api/restore', methods=['GET'])
def restore_from_db():
    """
    从数据库恢复到Excel：遍历所有预约，写入Excel
    前端下载后提示已恢复
    """
    import threading
    def _do_restore():
        try:
            from excel_sync import sync_all
            from models import get_db
            conn = get_db()
            rows = conn.execute(
                "SELECT a.*, p.name as patient_name FROM appointments a JOIN patients p ON a.patient_id=p.id"
            ).fetchall()
            conn.close()
            apts = [dict(r) for r in rows]
            sync_all(apts)
        except Exception:
            pass
    
    t = threading.Thread(target=_do_restore)
    t.start()
    return jsonify({'status': 'ok', 'message': '已从数据库恢复所有预约到Excel文件'})

# ===== 导出格式化 Excel（每月一个文件，打包 zip） =====

def _build_month_workbook(year, month, apts):
    """为单个月份构建一个 Workbook，样式与 2026预约表模板一致"""
    import datetime, calendar
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    WEEKDAY_NAMES = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']
    TIME_SLOTS = ['8:00-9:00', '9:00-10:00', '10:00-11:00', '11:00-12:00',
                  '12:00-13:00', '13:00-14:00', '14:00-15:00', '15:00-16:00',
                  '16:00-17:00', '17:00-18:00', '18:00-19:00', '19:00-20:00',
                  '20:00-21:00', '21:00-22:00']

    thin_side = Side(style='thin')
    thin_border = Border(left=thin_side, right=thin_side,
                         top=thin_side, bottom=thin_side)

    fill_title   = PatternFill(start_color='FF81D8CF', end_color='FF81D8CF', fill_type='solid')
    fill_white    = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
    fill_weekday  = PatternFill(start_color='FFE8D7CB', end_color='FFE8D7CB', fill_type='solid')
    fill_weekend  = PatternFill(start_color='FFFCDCA9', end_color='FFFCDCA9', fill_type='solid')

    font_title  = Font(name='Microsoft YaHei', bold=True,  size=14, color='FF000000')
    font_header = Font(name='Microsoft YaHei', bold=True,  size=14, color='FF000000')
    font_date   = Font(name='Microsoft YaHei', bold=True,  size=14, color='FF2F4154')
    font_cell   = Font(name='Microsoft YaHei', bold=False, size=14, color='FF000000')

    al_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    datamap = {}
    for a in apts:
        datamap[(a['day'], a['time_slot'])] = a

    days_in_month = calendar.monthrange(year, month)[1]

    def _day_groups():
        groups = []
        start = 1
        while start <= days_in_month:
            d = datetime.date(year, month, start)
            to_sunday = 6 - d.weekday()
            end = min(start + to_sunday, days_in_month)
            groups.append((start, end))
            start = end + 1
        return groups

    wb = Workbook()
    wb.remove(wb.active)

    for start_day, end_day in _day_groups():
        num_days = end_day - start_day + 1
        total_cols = 1 + num_days * 2
        sheet_name = f'{start_day}-{end_day}'
        ws = wb.create_sheet(title=sheet_name)
        ws.freeze_panes = 'B2'

        # ===== Row 1: 标题 =====
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=total_cols)
        c = ws.cell(row=1, column=1,
                    value=f'{year}年{month:02d}月预约表')
        c.font, c.fill, c.alignment, c.border = (
            font_title, fill_title, al_center, thin_border)
        ws.row_dimensions[1].height = 45

        # ===== Row 2-3: 时段 + 日期头 =====
        ws.merge_cells(start_row=2, start_column=1,
                       end_row=3, end_column=1)
        c = ws.cell(row=2, column=1, value='时段')
        c.font, c.fill, c.alignment, c.border = (
            font_header, fill_white, al_center, thin_border)
        c3 = ws.cell(row=3, column=1)
        c3.font, c3.fill, c3.border = font_header, fill_white, thin_border
        ws.row_dimensions[2].height = 45
        ws.row_dimensions[3].height = 45

        for i, day in enumerate(range(start_day, end_day + 1)):
            cn = 2 + i * 2
            cr = cn + 1

            dt = datetime.date(year, month, day)
            wd = WEEKDAY_NAMES[dt.weekday()]
            is_wknd = dt.weekday() >= 5
            day_fill = fill_weekend if is_wknd else fill_weekday

            ws.merge_cells(start_row=2, start_column=cn,
                           end_row=2, end_column=cr)
            c = ws.cell(row=2, column=cn,
                        value=f'{wd}\n{month:02d}-{day:02d}')
            c.font, c.alignment, c.border = font_date, al_center, thin_border
            for cc in range(cn, cr + 1):
                rc = ws.cell(row=2, column=cc)
                rc.fill, rc.border = day_fill, thin_border

            for col, label in [(cn, '预约'), (cr, '备注')]:
                c = ws.cell(row=3, column=col, value=label)
                c.font, c.fill, c.alignment, c.border = (
                    font_header, fill_white, al_center, thin_border)

        # ===== Rows 4-15: 时段 + 数据 =====
        for ri, ts in enumerate(TIME_SLOTS):
            row = 4 + ri
            ws.row_dimensions[row].height = 45

            c = ws.cell(row=row, column=1, value=ts)
            c.font, c.fill, c.alignment, c.border = (
                font_cell, fill_white, al_center, thin_border)

            for i, day in enumerate(range(start_day, end_day + 1)):
                cn = 2 + i * 2
                cr = cn + 1
                key = (day, ts)
                if key in datamap:
                    a = datamap[key]
                    nc = ws.cell(row=row, column=cn,
                                 value=a['patient_name'])
                    nc.font, nc.fill, nc.alignment, nc.border = (
                        font_cell, fill_white, al_center, thin_border)
                    parts = []
                    if a['project']:
                        parts.append(a['project'])
                    if a['amount']:
                        parts.append(str(a['amount']))
                    rc = ws.cell(row=row, column=cr,
                                 value='\n'.join(parts) if parts else None)
                    rc.font, rc.fill, rc.alignment, rc.border = (
                        font_cell, fill_white, al_center, thin_border)
                else:
                    for col in (cn, cr):
                        ec = ws.cell(row=row, column=col)
                        ec.font, ec.fill, ec.alignment, ec.border = (
                            font_cell, fill_white, al_center, thin_border)

        # ===== 列宽 =====
        ws.column_dimensions['A'].width = 14.625
        for i in range(1, total_cols + 1):
            cl = get_column_letter(i)
            if cl != 'A':
                ws.column_dimensions[cl].width = 12

    return wb


@app.route('/api/export-excel', methods=['GET'])
def export_excel():
    """导出格式化的 Excel 预约表 — 每月一个文件，打包为 zip 下载"""
    import io, time, zipfile
    from collections import defaultdict

    conn = get_db()
    appointments = conn.execute("""
        SELECT a.*, p.name as patient_name
        FROM appointments a
        LEFT JOIN patients p ON a.patient_id = p.id
        ORDER BY a.year, a.month, a.day, a.time_slot
    """).fetchall()
    conn.close()

    if not appointments:
        # 即使没有预约数据，也生成当前年份的空表
        by_month = {}
        years = {2026}
    else:
        by_month = defaultdict(list)
        for a in appointments:
            by_month[(a['year'], a['month'])].append(a)
        years = set(y for (y, m) in by_month.keys())

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for year in sorted(years):
            for month in range(1, 13):
                apts = by_month.get((year, month), [])
                wb = _build_month_workbook(year, month, apts)
                xlsx_buf = io.BytesIO()
                wb.save(xlsx_buf)
                wb.close()
                xlsx_buf.seek(0)
                fname = f'{year}年{month:02d}月预约表.xlsx'
                zf.writestr(fname, xlsx_buf.read())
                xlsx_buf.close()

    zip_buf.seek(0)
    zip_name = f'预约表导出_{time.strftime("%Y%m%d_%H%M%S")}.zip'
    return send_file(
        zip_buf,
        mimetype='application/zip',
        as_attachment=True,
        download_name=zip_name
    )

# ===== 患者管理 =====

@app.route('/api/patients', methods=['GET'])
def get_patients():
    search = request.args.get('search', '')
    
    conn = get_db()
    query = "SELECT * FROM patients WHERE 1=1"
    params = []
    
    if search:
        query += " AND name LIKE ?"
        params.append(f'%{search}%')
    
    query += " ORDER BY name"
    patients = conn.execute(query, params).fetchall()
    conn.close()
    
    return jsonify([dict(p) for p in patients])

@app.route('/api/patients', methods=['POST'])
def create_patient():
    data = request.json
    name = data.get('name', '').strip()
    
    if not name:
        return jsonify({'error': '姓名不能为空'}), 400
    
    conn = get_db()
    try:
        conn.execute("INSERT INTO patients (name, phone, age, gender) VALUES (?, ?, ?, ?)",
                     (name, data.get('phone', ''), data.get('age', ''), data.get('gender', '')))
        conn.commit()
        pid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        patient = conn.execute("SELECT * FROM patients WHERE id=?", (pid,)).fetchone()
        conn.close()
        return jsonify(dict(patient))
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 400

@app.route('/api/patients/<int:pid>', methods=['PUT'])
def update_patient(pid):
    data = request.json
    conn = get_db()
    updates = []
    params = []
    
    if 'name' in data:
        updates.append("name=?")
        params.append(data['name'].strip())
    if 'phone' in data:
        updates.append("phone=?")
        params.append(data['phone'])
    if 'age' in data:
        updates.append("age=?")
        params.append(data['age'])
    if 'gender' in data:
        updates.append("gender=?")
        params.append(data['gender'])
    if 'balance' in data:
        updates.append("balance=?")
        params.append(data['balance'])
    if 'has_recharge' in data:
        updates.append("has_recharge=?")
        params.append(data['has_recharge'])
    
    if updates:
        params.append(pid)
        conn.execute(f"UPDATE patients SET {','.join(updates)} WHERE id=?", params)
        conn.commit()
    
    patient = conn.execute("SELECT * FROM patients WHERE id=?", (pid,)).fetchone()
    conn.close()
    return jsonify(dict(patient))

@app.route('/api/patients/<int:pid>', methods=['DELETE'])
def delete_patient(pid):
    conn = get_db()
    # Sync: delete patient's appointments from Excel first
    apts = conn.execute("SELECT * FROM appointments WHERE patient_id=?", (pid,)).fetchall()
    for a in apts:
        sync_delete(a['year'], a['month'], a['day'], a['time_slot'])
    conn.execute("DELETE FROM appointments WHERE patient_id=?", (pid,))
    conn.execute("DELETE FROM consumption WHERE patient_id=?", (pid,))
    conn.execute("DELETE FROM patients WHERE id=?", (pid,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# ===== 预约管理 =====

@app.route('/api/schedule', methods=['GET'])
def get_schedule():
    import openpyxl, re, os
    month = request.args.get('month', '')
    if not month:
        return jsonify({'error': '请指定月份'}), 400
    try:
        month = int(month)
        if not (1 <= month <= 12):
            return jsonify({'error': '月份需在1-12之间'}), 400
    except ValueError:
        return jsonify({'error': '月份格式无效'}), 400
    
    try:
        base_dir = r'C:\Users\26259\Desktop\预约\2026预约表'
    except:
        base_dir = os.path.join(os.path.dirname(__file__), 'data', 'appointments')
    filename = None
    for f in os.listdir(base_dir):
        if f.endswith('.xlsx') and re.search(rf'{month}\u6708', f):
            filename = f
            break
    if not filename:
        return jsonify({'error': '未找到该月预约表'}), 404
    
    wb = openpyxl.load_workbook(os.path.join(base_dir, filename))
    appointments = []
    
    for sname in wb.sheetnames:
        ws = wb[sname]
        # Parse date headers from row 2
        dates = []
        for col in range(2, 30):
            val = ws.cell(row=2, column=col).value
            if val:
                m = re.search(r'(\d{2})-(\d{2})', str(val))
                if m:
                    dates.append({'col': col, 'day': int(m.group(2)), 'weekday': str(val).strip()})
        
        if not dates:
            continue
        
        # Parse rows 4-15
        for r in range(4, 18):
            time_slot = ws.cell(row=r, column=1).value
            if not time_slot:
                continue
            time_slot = str(time_slot).strip()
            
            for d in dates:
                name_val = ws.cell(row=r, column=d['col']).value
                remark_val = ws.cell(row=r, column=d['col'] + 1).value
                
                if not name_val:
                    continue
                
                name = str(name_val).strip()
                remark = str(remark_val).strip() if remark_val else ''
                
                # Parse amount from remark
                amount = 0
                amount_match = re.search(r'(\d+)$', remark)
                if amount_match:
                    amount = int(amount_match.group(1))
                
                project = re.sub(r'[\d\.]+[Hh]?', '', remark)
                project = re.sub(r'[Hh]+$', '', project).strip()
                if not project:
                    project = '治疗'
                
                appointments.append({
                    'id': len(appointments) + 1,
                    'month': month,
                    'day': d['day'],
                    'time_slot': time_slot,
                    'patient_name': name,
                    'project': project,
                    'amount': amount,
                    'remark': remark
                })
    
    wb.close()
    return jsonify(appointments)

@app.route('/api/appointments', methods=['GET'])
def get_appointments():
    year = request.args.get('year', '2026')
    month = request.args.get('month', '')
    patient_id = request.args.get('patient_id', '')
    
    conn = get_db()
    query = """
        SELECT a.*, p.name as patient_name
        FROM appointments a
        LEFT JOIN patients p ON a.patient_id = p.id
        WHERE 1=1
    """
    params = []
    
    if year:
        query += " AND a.year=?"
        params.append(int(year))
    if month:
        query += " AND a.month=?"
        params.append(int(month))
    if patient_id:
        query += " AND a.patient_id=?"
        params.append(int(patient_id))
    
    query += " ORDER BY a.month, a.day, a.time_slot"
    appointments = conn.execute(query, params).fetchall()
    conn.close()
    
    return jsonify([dict(a) for a in appointments])

@app.route('/api/sync-excel', methods=['POST'])
def sync_excel():
    import threading
    def _do_sync():
        conn = get_db()
        rows = conn.execute(
            "SELECT a.*, p.name as patient_name FROM appointments a JOIN patients p ON a.patient_id=p.id"
        ).fetchall()
        conn.close()
        apts = [dict(r) for r in rows]
        sync_all(apts)  # 异步，会在后台线程中执行
    
    t = threading.Thread(target=_do_sync)
    t.start()
    return jsonify({'status': 'queued', 'message': 'Excel 同步任务已加入队列，将在后台执行'})

@app.route('/api/sync-status', methods=['GET'])
def sync_status():
    from excel_sync import get_queue_status
    return jsonify(get_queue_status())

@app.route('/api/appointments', methods=['POST'])
def create_appointment():
    data = request.json
    patient_id = data.get('patient_id')
    year = data.get('year', 2026)
    month = data.get('month')
    day = data.get('day')
    time_slot = data.get('time_slot', '')
    project = data.get('project', '治疗')
    amount = data.get('amount', 0)
    remark = data.get('remark', '')
    recharge_amt = data.get('recharge_amount', 0) or 0
    
    if not patient_id or not month or not day:
        return jsonify({'error': '缺少必要字段'}), 400
    
    try: _validate_date(int(month), int(day))
    except ValueError as e: return jsonify({'error': str(e)}), 400
    
    conn = get_db()
    patient = conn.execute("SELECT id, name FROM patients WHERE id=?", (patient_id,)).fetchone()
    if not patient:
        conn.close()
        return jsonify({'error': '患者不存在'}), 400
    
    conn.execute(
        "INSERT INTO appointments (patient_id, year, month, day, time_slot, project, amount, remark, recharge_amount) VALUES (?,?,?,?,?,?,?,?,?)",
        (patient_id, year, month, day, time_slot, project, amount, remark, recharge_amt)
    )
    conn.commit()
    aid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    
    # 充值：增加患者余额，自动扣本次消费
    if recharge_amt > 0:
        conn.execute("UPDATE patients SET balance = balance + ?, has_recharge = 1 WHERE id=?", (recharge_amt, patient_id))
        conn.commit()
    # 有余额时自动扣减本次金额
    _deduct_balance(conn, patient_id, amount)
    conn.commit()
    
    linked_con_id = None
    # 如果请求中包含 sync_to_con，创建关联的耗卡
    if data.get('sync_to_con'):
        conn.execute(
            "INSERT INTO consumption (patient_id, month, day, project, amount, time_slot, sync_to_apt, recharge_amount) VALUES (?,?,?,?,?,?,?,?)",
            (patient_id, month, day, project, amount, time_slot, 1, recharge_amt)
        )
        con_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        # 建立关联
        conn.execute(
            "INSERT INTO appointment_consumption_link (appointment_id, consumption_id) VALUES (?, ?)",
            (aid, con_id)
        )
        conn.commit()
        linked_con_id = con_id
    
    appointment = conn.execute(
        "SELECT a.*, p.name as patient_name FROM appointments a LEFT JOIN patients p ON a.patient_id=p.id WHERE a.id=?",
        (aid,)
    ).fetchone()
    conn.close()
    apt = dict(appointment) if appointment else {'patient_name': patient['name'], 'patient_id': patient_id, 'year': year, 'month': month, 'day': day, 'time_slot': time_slot, 'project': project, 'amount': amount, 'remark': remark, 'id': aid}
    if linked_con_id:
        apt['linked_con_id'] = linked_con_id
    try:
        sync_upsert(apt.get('year'), apt.get('month'), apt.get('day'), apt.get('time_slot',''),
                    patient['name'], apt.get('project', ''), apt.get('amount', 0), apt.get('remark', ''))
    except Exception:
        pass
    return jsonify(apt)

@app.route('/api/appointments/<int:aid>', methods=['PUT'])
def update_appointment(aid):
    data = request.json
    conn = get_db()
    
    old = conn.execute("SELECT * FROM appointments WHERE id=?", (aid,)).fetchone()
    if not old:
        conn.close()
        return jsonify({'error': '预约不存在'}), 404
    old = dict(old)
    
    updates = []
    params = []
    fields = ['patient_id', 'year', 'month', 'day', 'time_slot', 'project', 'amount', 'remark', 'recharge_amount']
    
    for f in fields:
        if f in data:
            updates.append(f"{f}=?")
            params.append(data[f])
    
    if updates:
        params.append(aid)
        conn.execute(f"UPDATE appointments SET {','.join(updates)} WHERE id=?", params)
        conn.commit()
    
    # 充值金额变更时，重算患者余额
    if 'recharge_amount' in data or 'amount' in data:
        pt_id = data.get('patient_id', old['patient_id'])
        total_recharge = conn.execute(
            "SELECT COALESCE(SUM(recharge_amount),0) FROM appointments WHERE patient_id=?", (pt_id,)
        ).fetchone()[0]
        total_spent = conn.execute(
            "SELECT COALESCE(SUM(amount),0) FROM appointments WHERE patient_id=?", (pt_id,)
        ).fetchone()[0]
        conn.execute("UPDATE patients SET balance=MAX(0,?), has_recharge=CASE WHEN ?>0 THEN 1 ELSE 0 END WHERE id=?",
                     (total_recharge - total_spent, total_recharge, pt_id))
        conn.commit()
    
    appointment = conn.execute(
        "SELECT a.*, p.name as patient_name FROM appointments a LEFT JOIN patients p ON a.patient_id=p.id WHERE a.id=?",
        (aid,)
    ).fetchone()
    
    if not appointment:
        conn.close()
        return jsonify({'error': '获取预约失败'}), 500

    apt = dict(appointment)
    patient_name = apt.get('patient_name') or ''
    try:
        old_cell_changed = (old['year'] != apt['year'] or old['month'] != apt['month'] or
                            old['day'] != apt['day'] or apt['time_slot'] != old['time_slot'])
        if old_cell_changed:
            sync_delete(old['year'], old['month'], old['day'], old['time_slot'])
        sync_upsert(apt['year'], apt['month'], apt['day'], apt['time_slot'],
                    patient_name, apt.get('project', ''), apt.get('amount', 0), apt.get('remark', ''))
    except Exception:
        pass
    # 使用关联表查找关联的耗卡，而不是 remark 中的 #id
    linked_cons = conn.execute(
        "SELECT consumption_id FROM appointment_consumption_link WHERE appointment_id=?",
        (aid,)
    ).fetchall()
    
    if linked_cons:
        # Update via link table (precise)
        for link in linked_cons:
            cid = link['consumption_id']
            conn.execute(
                "UPDATE consumption SET month=?, day=?, project=?, amount=?, time_slot=?, sync_to_apt=1, recharge_amount=? WHERE id=?",
                (apt['month'], apt['day'], apt['project'], apt['amount'], apt['time_slot'], apt.get('recharge_amount', 0), cid)
            )
    else:
        # Fallback: match by patient + month + day + time_slot (natural key)
        conn.execute(
            "UPDATE consumption SET month=?, day=?, project=?, amount=?, time_slot=?, sync_to_apt=1, recharge_amount=? WHERE patient_id=? AND month=? AND day=? AND time_slot=?",
            (apt['month'], apt['day'], apt['project'], apt['amount'], apt['time_slot'], apt.get('recharge_amount', 0),
             old['patient_id'], old['month'], old['day'], old['time_slot'])
        )
    # 关联consumption更新后，也从consumption表重算balance/has_recharge
    # 确保与前端数据源（consumption）一致
    pt_id = apt.get('patient_id')
    if pt_id:
        total_recharge = conn.execute(
            "SELECT COALESCE(SUM(recharge_amount),0) FROM consumption WHERE patient_id=?", (pt_id,)
        ).fetchone()[0]
        total_spent = conn.execute(
            "SELECT COALESCE(SUM(amount),0) FROM consumption WHERE patient_id=?", (pt_id,)
        ).fetchone()[0]
        conn.execute(
            "UPDATE patients SET balance=MAX(0,?), has_recharge=CASE WHEN ?>0 THEN 1 ELSE 0 END WHERE id=?",
            (total_recharge - total_spent, total_recharge, pt_id)
        )
    conn.commit()
    conn.close()
    return jsonify(apt)

@app.route('/api/appointments/<int:aid>', methods=['DELETE'])
def delete_appointment(aid):
    conn = get_db()
    apt = conn.execute("SELECT a.*, p.name as patient_name FROM appointments a LEFT JOIN patients p ON a.patient_id=p.id WHERE a.id=?", (aid,)).fetchone()
    if apt:
        # 删除关联的同步耗卡（通过关联表）
        linked_cons = conn.execute(
            "SELECT consumption_id FROM appointment_consumption_link WHERE appointment_id=?",
            (aid,)
        ).fetchall()
        for link in linked_cons:
            conn.execute("DELETE FROM consumption WHERE id=?", (link['consumption_id'],))
        # 同时清理关联表记录
        conn.execute("DELETE FROM appointment_consumption_link WHERE appointment_id=?", (aid,))
    conn.execute("DELETE FROM appointments WHERE id=?", (aid,))
    conn.commit()
    conn.close()
    if apt:
        try:
            sync_delete(apt['year'], apt['month'], apt['day'], apt['time_slot'])
        except Exception:
            pass
    return jsonify({'success': True})

# ===== 耗卡管理 =====

@app.route('/api/consumption', methods=['GET'])
def get_consumption():
    patient_id = request.args.get('patient_id', '')
    month = request.args.get('month', '')
    
    conn = get_db()
    query = """
        SELECT c.*, p.name as patient_name
        FROM consumption c
        LEFT JOIN patients p ON c.patient_id = p.id
        WHERE 1=1
    """
    params = []
    
    if patient_id:
        query += " AND c.patient_id=?"
        params.append(int(patient_id))
    if month:
        query += " AND c.month=?"
        params.append(int(month))
    
    query += " ORDER BY c.month, c.day, c.id"
    records = conn.execute(query, params).fetchall()
    conn.close()
    
    return jsonify([dict(r) for r in records])

@app.route('/api/consumption', methods=['POST'])
def create_consumption():
    data = request.json
    patient_id = data.get('patient_id')
    month = data.get('month')
    day = data.get('day')
    project = data.get('project', '治疗')
    amount = data.get('amount', 0)
    time_slot = data.get('time_slot', '')
    sync_to_apt = data.get('sync_to_apt', 0)
    recharge_amt = data.get('recharge_amount', 0) or 0
    
    if not patient_id or not month or not day:
        return jsonify({'error': '缺少必要字段'}), 400
    
    try: _validate_date(int(month), int(day))
    except ValueError as e: return jsonify({'error': str(e)}), 400
    
    conn = get_db()
    conn.execute(
        "INSERT INTO consumption (patient_id, month, day, project, amount, time_slot, sync_to_apt, recharge_amount) VALUES (?,?,?,?,?,?,?,?)",
        (patient_id, month, day, project, amount, time_slot, sync_to_apt, recharge_amt)
    )
    conn.commit()
    cid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    
    # 充值：增加患者余额，自动扣本次消费
    if recharge_amt > 0:
        conn.execute("UPDATE patients SET balance = balance + ?, has_recharge = 1 WHERE id=?", (recharge_amt, patient_id))
        conn.commit()
    # 有余额时自动扣减本次金额
    _deduct_balance(conn, patient_id, amount)
    conn.commit()
    
    record = conn.execute(
        "SELECT c.*, p.name as patient_name FROM consumption c LEFT JOIN patients p ON c.patient_id=p.id WHERE c.id=?",
        (cid,)
    ).fetchone()
    conn.close()
    return jsonify(dict(record) if record else {})

@app.route('/api/consumption/<int:cid>', methods=['PUT'])
def update_consumption(cid):
    data = request.json
    conn = get_db()
    updates = []
    params = []
    fields = ['patient_id', 'month', 'day', 'project', 'amount', 'time_slot', 'sync_to_apt', 'recharge_amount']
    
    for f in fields:
        if f in data:
            updates.append(f"{f}=?")
            params.append(data[f])
    
    if updates:
        params.append(cid)
        conn.execute(f"UPDATE consumption SET {','.join(updates)} WHERE id=?", params)
        conn.commit()
    
    # After update, recalculate patient balance and has_recharge from consumption
    record = conn.execute("SELECT * FROM consumption WHERE id=?", (cid,)).fetchone()
    if record:
        patient_id = record['patient_id']
        total_recharge = conn.execute(
            "SELECT COALESCE(SUM(recharge_amount),0) FROM consumption WHERE patient_id=?", (patient_id,)
        ).fetchone()[0]
        total_spent = conn.execute(
            "SELECT COALESCE(SUM(amount),0) FROM consumption WHERE patient_id=?", (patient_id,)
        ).fetchone()[0]
        conn.execute(
            "UPDATE patients SET balance=MAX(0,?), has_recharge=CASE WHEN ?>0 THEN 1 ELSE 0 END WHERE id=?",
            (total_recharge - total_spent, total_recharge, patient_id)
        )
        conn.commit()
    
    record = conn.execute(
        "SELECT c.*, p.name as patient_name FROM consumption c LEFT JOIN patients p ON c.patient_id=p.id WHERE c.id=?",
        (cid,)
    ).fetchone()
    conn.close()
    if not record:
        return jsonify({'error': '耗卡记录不存在'}), 404
    return jsonify(dict(record))

@app.route('/api/consumption/<int:cid>', methods=['DELETE'])
def delete_consumption(cid):
    conn = get_db()
    conn.execute("DELETE FROM consumption WHERE id=?", (cid,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# ===== 批量导入消费记录 =====

@app.route('/api/consumption/batch', methods=['POST'])
def batch_create_consumption():
    data = request.json
    records = data.get('records', [])
    if len(records) > _BATCH_LIMIT:
        return jsonify({'error': f'单次最多{_BATCH_LIMIT}条'}), 400
    
    conn = get_db()
    count = 0
    for r in records:
        ts = r.get('time_slot', '')
        # Exact match (including time_slot) → skip
        if conn.execute(
            "SELECT 1 FROM consumption WHERE patient_id=? AND month=? AND day=? AND project=? AND amount=? AND time_slot=?",
            (r['patient_id'], r['month'], r['day'], r['project'], r['amount'], ts)
        ).fetchone():
            continue
        if ts:
            # Match same record with empty time_slot → backfill
            existing = conn.execute(
                "SELECT id FROM consumption WHERE patient_id=? AND month=? AND day=? AND project=? AND amount=? AND (time_slot IS NULL OR time_slot='')",
                (r['patient_id'], r['month'], r['day'], r['project'], r['amount'])
            ).fetchone()
            if existing:
                conn.execute("UPDATE consumption SET time_slot=?, sync_to_apt=1 WHERE id=?", (ts, existing['id']))
                count += 1
                continue
        conn.execute(
            "INSERT INTO consumption (patient_id, month, day, project, amount, time_slot, sync_to_apt) VALUES (?,?,?,?,?,?,?)",
            (r['patient_id'], r['month'], r['day'], r['project'], r['amount'], ts, 1)
        )
        count += 1
    conn.commit()
    conn.close()
    return jsonify({'count': count})

# ===== 治疗项目管理 =====

@app.route('/api/projects', methods=['GET'])
def get_projects():
    conn = get_db()
    projects = conn.execute("SELECT * FROM treatment_projects ORDER BY sort_order, id").fetchall()
    conn.close()
    return jsonify([dict(p) for p in projects])

@app.route('/api/projects', methods=['POST'])
def create_project():
    data = request.json
    name = data.get('name', '').strip()
    default_price = data.get('default_price', 0)
    if not name:
        return jsonify({'error': '项目名称不能为空'}), 400
    conn = get_db()
    try:
        conn.execute("INSERT INTO treatment_projects (name, default_price) VALUES (?,?)", (name, default_price))
        conn.commit()
        pid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        project = conn.execute("SELECT * FROM treatment_projects WHERE id=?", (pid,)).fetchone()
        conn.close()
        return jsonify(dict(project))
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 400

@app.route('/api/projects/<int:pid>', methods=['PUT'])
def update_project(pid):
    data = request.json
    conn = get_db()
    updates = []
    params = []
    fields = ['name', 'default_price', 'sort_order']
    for f in fields:
        if f in data:
            updates.append(f"{f}=?")
            params.append(data[f])
    if updates:
        params.append(pid)
        conn.execute(f"UPDATE treatment_projects SET {','.join(updates)} WHERE id=?", params)
        conn.commit()
    project = conn.execute("SELECT * FROM treatment_projects WHERE id=?", (pid,)).fetchone()
    conn.close()
    return jsonify(dict(project))

@app.route('/api/projects/<int:pid>', methods=['DELETE'])
def delete_project(pid):
    conn = get_db()
    conn.execute("DELETE FROM treatment_projects WHERE id=?", (pid,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# ===== 统计 =====

@app.route('/api/stats', methods=['GET'])
def get_stats():
    year = request.args.get('year', '2026')
    conn = get_db()
    
    # 月度营收
    monthly = conn.execute("""
        SELECT month, SUM(amount) as total, COUNT(*) as count
        FROM appointments
        WHERE year=?
        GROUP BY month
        ORDER BY month
    """, (year,)).fetchall()
    
    # 患者统计
    total_patients = conn.execute("SELECT COUNT(*) FROM patients").fetchone()[0]
    
    # 患者到诊统计
    patient_visits = conn.execute("""
        SELECT p.name, COUNT(a.id) as visit_count, SUM(a.amount) as total_amount
        FROM patients p
        LEFT JOIN appointments a ON p.id = a.patient_id AND a.year=?
        GROUP BY p.id
        HAVING visit_count > 0
        ORDER BY visit_count DESC
    """, (year,)).fetchall()
    
    # 项目统计
    project_stats = conn.execute("""
        SELECT project, COUNT(*) as count, SUM(amount) as total_amount
        FROM appointments
        WHERE year=?
        GROUP BY project
        ORDER BY count DESC
    """, (year,)).fetchall()
    
    # 年度总额
    total_revenue = conn.execute("""
        SELECT COALESCE(SUM(amount),0) FROM appointments WHERE year=?
    """, (year,)).fetchone()[0]
    
    # 月均营收
    month_count = len(monthly)
    avg_revenue = total_revenue / month_count if month_count > 0 else 0
    
    conn.close()
    
    return jsonify({
        'year': year,
        'total_revenue': total_revenue,
        'avg_revenue': round(avg_revenue, 2),
        'total_patients': total_patients,
        'monthly': [dict(m) for m in monthly],
        'patient_visits': [dict(p) for p in patient_visits],
        'project_stats': [dict(p) for p in project_stats]
    })

# ===== 导入数据 =====

@app.route('/api/import', methods=['POST'])
def import_data():
    try:
        stats = import_from_excel()
        return jsonify(stats)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ===== 操作历史 =====
import json as _json

@app.route('/api/history', methods=['GET'])
def get_history():
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM operation_history ORDER BY created_at DESC LIMIT 200"
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/history', methods=['POST'])
def add_history():
    data = request.json
    conn = get_db()
    data_val = data.get('data', {})
    if isinstance(data_val, str):
        data_val_str = data_val  # already JSON string from frontend
    else:
        data_val_str = _json.dumps(data_val, ensure_ascii=False)
    conn.execute(
        "INSERT INTO operation_history (action_type, target_type, target_id, data, description) VALUES (?,?,?,?,?)",
        (data['action_type'], data['target_type'], data.get('target_id'),
         data_val_str, data.get('description', ''))
    )
    conn.commit()
    hid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    record = conn.execute("SELECT * FROM operation_history WHERE id=?", (hid,)).fetchone()
    conn.close()
    return jsonify(dict(record))

@app.route('/api/history/<int:hid>/restore', methods=['POST'])
def restore_history(hid):
    conn = get_db()
    row = conn.execute("SELECT * FROM operation_history WHERE id=?", (hid,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': '记录不存在'}), 404

    # Find all entries after this one (newer) to reverse first
    later = conn.execute(
        "SELECT * FROM operation_history WHERE id > ? ORDER BY id DESC",
        (hid,)
    ).fetchall()

    entries_to_reverse = list(later) + [row]
    errors = []
    from excel_sync import sync_delete, sync_upsert
    id_map = {}  # 记录重建后的新 ID 映射

    for entry in entries_to_reverse:
        try:
            data = _json.loads(entry['data'])
            act = entry['action_type']
            ttype = entry['target_type']
            tid = entry['target_id']
            real_id = id_map.get(('appointment', tid), id_map.get(('consumption', tid), id_map.get(('patient', tid), tid)))

            if act == 'add':
                # Reverse: delete the record
                if ttype == 'appointment':
                    remark = data.get('remark', '')
                    m = re.search(r'#(\d+)$', remark) if remark else None
                    if m:
                        con_real = id_map.get(('consumption', int(m.group(1))), int(m.group(1)))
                        conn.execute("DELETE FROM consumption WHERE id=?", (con_real,))
                    conn.execute("DELETE FROM appointments WHERE id=?", (real_id,))
                    # 重算患者 balance 和 has_recharge
                    pt_id = data.get('patient_id')
                    if pt_id:
                        total_recharge = conn.execute(
                            "SELECT COALESCE(SUM(recharge_amount),0) FROM appointments WHERE patient_id=?", (pt_id,)
                        ).fetchone()[0]
                        total_spent = conn.execute(
                            "SELECT COALESCE(SUM(amount),0) FROM appointments WHERE patient_id=?", (pt_id,)
                        ).fetchone()[0]
                        conn.execute(
                            "UPDATE patients SET balance=MAX(0,?), has_recharge=CASE WHEN ?>0 THEN 1 ELSE 0 END WHERE id=?",
                            (total_recharge - total_spent, total_recharge, pt_id)
                        )
                    sync_delete(data.get('year'), data.get('month'), data.get('day'), data.get('time_slot',''))
                elif ttype == 'consumption':
                    # 删前先获取 patient_id
                    c_rec = conn.execute("SELECT patient_id FROM consumption WHERE id=?", (real_id,)).fetchone()
                    conn.execute("DELETE FROM consumption WHERE id=?", (real_id,))
                    if c_rec:
                        c_pt_id = c_rec['patient_id']
                        total_recharge = conn.execute(
                            "SELECT COALESCE(SUM(recharge_amount),0) FROM consumption WHERE patient_id=?", (c_pt_id,)
                        ).fetchone()[0]
                        total_spent = conn.execute(
                            "SELECT COALESCE(SUM(amount),0) FROM consumption WHERE patient_id=?", (c_pt_id,)
                        ).fetchone()[0]
                        conn.execute(
                            "UPDATE patients SET balance=MAX(0,?), has_recharge=CASE WHEN ?>0 THEN 1 ELSE 0 END WHERE id=?",
                            (total_recharge - total_spent, total_recharge, c_pt_id)
                        )
                elif ttype == 'patient':
                    apts = conn.execute("SELECT * FROM appointments WHERE patient_id=?", (real_id,)).fetchall()
                    for a in apts:
                        sync_delete(a['year'], a['month'], a['day'], a['time_slot'])
                    conn.execute("DELETE FROM appointments WHERE patient_id=?", (real_id,))
                    conn.execute("DELETE FROM consumption WHERE patient_id=?", (real_id,))
                    conn.execute("DELETE FROM patients WHERE id=?", (real_id,))
            elif act == 'delete':
                # Reverse: recreate the record
                if ttype == 'appointment':
                    conn.execute(
                        "INSERT INTO appointments (patient_id, year, month, day, time_slot, project, amount, remark, recharge_amount) VALUES (?,?,?,?,?,?,?,?,?)",
                        (data.get('patient_id'), data.get('year'), data.get('month'), data.get('day'),
                         data.get('time_slot',''), data.get('project',''), data.get('amount',0), data.get('remark',''), data.get('recharge_amount', 0) or 0)
                    )
                    new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
                    id_map[('appointment', tid)] = new_id
                    # 如果原预约有关联耗卡，一并恢复
                    remark = data.get('remark', '')
                    if remark and re.search(r'#(\d+)$', remark):
                        m = re.search(r'#(\d+)$', remark)
                        orig_con_id = int(m.group(1))
                        conn.execute(
                            "INSERT INTO consumption (patient_id, month, day, project, amount, time_slot, sync_to_apt) VALUES (?,?,?,?,?,?,?)",
                            (data.get('patient_id'), data.get('month'), data.get('day'),
                             data.get('project',''), data.get('amount',0), data.get('time_slot',''), 1)
                        )
                        new_con_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
                        id_map[('consumption', orig_con_id)] = new_con_id
                        # 更新预约 remark 指向新耗卡 ID
                        new_remark = re.sub(r'#(\d+)$', f'#{new_con_id}', remark)
                        conn.execute("UPDATE appointments SET remark=? WHERE id=?", (new_remark, new_id))
                    sync_upsert(data.get('year'), data.get('month'), data.get('day'), data.get('time_slot',''),
                               data.get('patient_name',''), data.get('project',''), data.get('amount',0), data.get('remark',''))
                    # 重算患者 balance 和 has_recharge
                    pt_id = data.get('patient_id')
                    if pt_id:
                        total_recharge = conn.execute(
                            "SELECT COALESCE(SUM(recharge_amount),0) FROM appointments WHERE patient_id=?", (pt_id,)
                        ).fetchone()[0]
                        total_spent = conn.execute(
                            "SELECT COALESCE(SUM(amount),0) FROM appointments WHERE patient_id=?", (pt_id,)
                        ).fetchone()[0]
                        conn.execute(
                            "UPDATE patients SET balance=MAX(0,?), has_recharge=CASE WHEN ?>0 THEN 1 ELSE 0 END WHERE id=?",
                            (total_recharge - total_spent, total_recharge, pt_id)
                        )
                elif ttype == 'consumption':
                    conn.execute(
                        "INSERT INTO consumption (patient_id, month, day, project, amount, time_slot, sync_to_apt, recharge_amount) VALUES (?,?,?,?,?,?,?,?)",
                        (data.get('patient_id'), data.get('month'), data.get('day'),
                         data.get('project',''), data.get('amount',0), data.get('time_slot',''),
                         data.get('sync_to_apt', 0), data.get('recharge_amount', 0) or 0)
                    )
                    new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
                    id_map[('consumption', tid)] = new_id
                    # 重算患者 balance 和 has_recharge
                    c_pt_id = data.get('patient_id')
                    if c_pt_id:
                        total_recharge = conn.execute(
                            "SELECT COALESCE(SUM(recharge_amount),0) FROM consumption WHERE patient_id=?", (c_pt_id,)
                        ).fetchone()[0]
                        total_spent = conn.execute(
                            "SELECT COALESCE(SUM(amount),0) FROM consumption WHERE patient_id=?", (c_pt_id,)
                        ).fetchone()[0]
                        conn.execute(
                            "UPDATE patients SET balance=MAX(0,?), has_recharge=CASE WHEN ?>0 THEN 1 ELSE 0 END WHERE id=?",
                            (total_recharge - total_spent, total_recharge, c_pt_id)
                        )
                elif ttype == 'patient':
                    conn.execute(
                        "INSERT INTO patients (name, phone, age, gender) VALUES (?,?,?,?)",
                        (data.get('name'), data.get('phone',''), data.get('age',''), data.get('gender',''))
                    )
                    new_pid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
                    id_map[('patient', tid)] = new_pid
                    # 重建关联的预约和耗卡
                    for apt in data.get('appointments', []):
                        conn.execute(
                            "INSERT INTO appointments (patient_id, year, month, day, time_slot, project, amount, remark) VALUES (?,?,?,?,?,?,?,?)",
                            (new_pid, apt.get('year', 2026), apt.get('month'), apt.get('day'),
                             apt.get('time_slot',''), apt.get('project',''), apt.get('amount',0), apt.get('remark',''))
                        )
                        new_apt_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
                        id_map[('appointment', apt['id'])] = new_apt_id
                    for con in data.get('consumptions', []):
                        conn.execute(
                            "INSERT INTO consumption (patient_id, month, day, project, amount, time_slot, sync_to_apt) VALUES (?,?,?,?,?,?,?)",
                            (new_pid, con.get('month'), con.get('day'),
                             con.get('project',''), con.get('amount',0), con.get('time_slot',''), con.get('sync_to_apt', 0))
                        )
                        new_con_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
                        id_map[('consumption', con['id'])] = new_con_id
                    # 修复所有预约 remark 中的耗卡 ID
                    for apt in data.get('appointments', []):
                        old_remark = apt.get('remark', '')
                        rm = re.search(r'#(\d+)$', old_remark) if old_remark else None
                        if rm:
                            old_cid = int(rm.group(1))
                            new_cid = id_map.get(('consumption', old_cid))
                            if new_cid:
                                new_remark = re.sub(r'#(\d+)$', f'#{new_cid}', old_remark)
                                new_aid = id_map.get(('appointment', apt['id']))
                                if new_aid:
                                    conn.execute("UPDATE appointments SET remark=? WHERE id=?", (new_remark, new_aid))
            elif act == 'update':
                # Reverse: restore old data
                old = data.get('old', {})
                newd = data.get('new', {})
                if ttype == 'appointment':
                    # 恢复预约旧值
                    old_remark = old.get('remark', '')
                    old_recharge = old.get('recharge_amount', 0) or 0
                    conn.execute(
                        "UPDATE appointments SET patient_id=?, month=?, day=?, time_slot=?, project=?, amount=?, remark=?, recharge_amount=? WHERE id=?",
                        (old.get('patient_id'), old.get('month'), old.get('day'),
                         old.get('time_slot',''), old.get('project',''), old.get('amount',0), old_remark, old_recharge, tid)
                    )
                    # 恢复关联耗卡
                    m = re.search(r'#(\d+)$', old_remark) if old_remark else None
                    if m:
                        conn.execute(
                            "UPDATE consumption SET patient_id=?, month=?, day=?, project=?, amount=?, time_slot=?, recharge_amount=? WHERE id=?",
                            (old.get('patient_id'), old.get('month'), old.get('day'),
                             old.get('project',''), old.get('amount',0), old.get('time_slot',''), old_recharge, int(m.group(1)))
                        )
                    # 重算患者 balance 和 has_recharge
                    pt_id = old.get('patient_id')
                    if pt_id:
                        total_recharge = conn.execute(
                            "SELECT COALESCE(SUM(recharge_amount),0) FROM appointments WHERE patient_id=?", (pt_id,)
                        ).fetchone()[0]
                        total_spent = conn.execute(
                            "SELECT COALESCE(SUM(amount),0) FROM appointments WHERE patient_id=?", (pt_id,)
                        ).fetchone()[0]
                        conn.execute(
                            "UPDATE patients SET balance=MAX(0,?), has_recharge=CASE WHEN ?>0 THEN 1 ELSE 0 END WHERE id=?",
                            (total_recharge - total_spent, total_recharge, pt_id)
                        )
                    # 同步Excel
                    sync_delete(newd.get('year'), newd.get('month'), newd.get('day'), newd.get('time_slot',''))
                    sync_upsert(old.get('year'), old.get('month'), old.get('day'), old.get('time_slot',''),
                               old.get('patient_name',''), old.get('project',''), old.get('amount',0), old_remark)
        except Exception as e:
            errors.append({'entry': entry['id'], 'error': str(e)})
            conn.rollback()
            conn.close()
            return jsonify({'error': '恢复失败', 'details': errors}), 500

    conn.commit()
    conn.close()

    # Also delete the restored entries from history
    conn2 = get_db()
    conn2.execute("DELETE FROM operation_history WHERE id >= ?", (hid,))
    conn2.commit()
    conn2.close()

    return jsonify({'success': True, 'reversed': len(entries_to_reverse), 'errors': errors})

if __name__ == '__main__':
    import socket
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    s.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
    try:
        s.bind(('0.0.0.0', 8080))
    except OSError:
        pass
    s.close()
    app.run(host='0.0.0.0', port=8080, debug=False)
