import openpyxl, sqlite3, os, re, glob

DB_PATH = '/root/clinic-app/data/clinic.db'
EXCEL_DIR = '/tmp/excel_data'

def main():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    patients_cache = {}
    c.execute('SELECT id, name FROM patients')
    for row in c.fetchall():
        patients_cache[row[1]] = row[0]

    files = sorted(glob.glob(os.path.join(EXCEL_DIR, '*.xlsx')))
    total_apt = 0
    total_con = 0
    total_pat = 0

    for fpath in files:
        fname = os.path.basename(fpath)
        wb = openpyxl.load_workbook(fpath, data_only=True)

        title = str(wb.active.cell(1, 1).value or '')
        month_match = re.search(r'年(\d{1,2})月', title)
        if not month_match:
            print(f'Skip {fname}: cannot parse month')
            continue
        month = int(month_match.group(1))
        print(f'Processing {fname} (month={month})...')

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            dates = {}
            row2 = [ws.cell(2, col).value for col in range(1, ws.max_column + 1)]
            for col_idx in range(1, len(row2)):
                val = row2[col_idx - 1]
                if val:
                    date_match = re.search(r'(\d{2})-(\d{2})', str(val))
                    if date_match:
                        day = int(date_match.group(2))
                        dates[col_idx] = day

            for row_idx in range(4, ws.max_row + 1):
                time_slot = ws.cell(row_idx, 1).value
                if not time_slot:
                    continue
                time_slot = str(time_slot).strip()

                for apt_col, day in dates.items():
                    note_col = apt_col + 1
                    if note_col > ws.max_column:
                        continue

                    patient_name = ws.cell(row_idx, apt_col).value
                    note_raw = ws.cell(row_idx, note_col).value

                    if not patient_name or not str(patient_name).strip():
                        continue

                    patient_name = str(patient_name).strip()
                    remark = str(note_raw).strip() if note_raw else ''

                    remark_clean = remark.replace('\n', ' ')
                    amount = 0
                    amount_match = re.search(r'(\d+)$', remark_clean)
                    if amount_match:
                        amount = int(amount_match.group(1))

                    project = re.sub(r'[\d\.]+[Hh]?', '', remark_clean)
                    project = re.sub(r'[Hh]+$', '', project).strip()
                    if not project:
                        project = '治疗'

                    if patient_name not in patients_cache:
                        c.execute('INSERT OR IGNORE INTO patients (name, status) VALUES (?, ?)',
                                  (patient_name, '未成交'))
                        if c.rowcount > 0:
                            total_pat += 1
                        c.execute('SELECT id FROM patients WHERE name=?', (patient_name,))
                        patients_cache[patient_name] = c.fetchone()[0]

                    pid = patients_cache[patient_name]

                    c.execute('''SELECT id FROM appointments
                                 WHERE patient_id=? AND year=2026 AND month=? AND day=? AND time_slot=?''',
                              (pid, month, day, time_slot))
                    existing = c.fetchone()
                    if existing:
                        apt_id = existing[0]
                    else:
                        c.execute('''INSERT INTO appointments (patient_id, year, month, day, time_slot, project, amount, remark)
                                     VALUES (?, 2026, ?, ?, ?, ?, ?, ?)''',
                                  (pid, month, day, time_slot, project, amount, remark))
                        apt_id = c.lastrowid
                        total_apt += 1

                    if amount > 0:
                        c.execute('''INSERT INTO consumption (patient_id, month, day, time_slot, project, amount, sync_to_apt)
                                     VALUES (?, ?, ?, ?, ?, ?, 1)''',
                                  (pid, month, day, time_slot, project, amount))
                        con_id = c.lastrowid
                        total_con += 1

                        c.execute('INSERT OR IGNORE INTO appointment_consumption_link (appointment_id, consumption_id) VALUES (?, ?)',
                                  (apt_id, con_id))

    conn.commit()
    conn.close()
    print(f'\nDone! Patients: {total_pat}, Appointments: {total_apt}, Consumptions: {total_con}')

if __name__ == '__main__':
    main()
